"""Build the BOQ workbook with openpyxl.

Sheets: BOQ (CPWD/DSR measurement-sheet columns, live Amount/SUM formulas so
the client can tweak rates), Bar Bending Schedule, and Assumptions.
See project.md section 8.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLUMNS = ["Item", "Description", "No.", "L (m)", "B (m)", "D/H (m)",
           "Quantity", "Unit", "Rate (INR)", "Amount (INR)"]


def _style_header(ws, row):
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = BORDER


def build_workbook(project: dict[str, Any], boq: dict[str, Any]) -> bytes:
    wb = Workbook()
    _boq_sheet(wb, project, boq)
    _bbs_sheet(wb, boq)
    _assumptions_sheet(wb, project, boq)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _boq_sheet(wb, project, boq):
    ws = wb.active
    ws.title = "BOQ"

    ws.append([f"Bill of Quantities — {project.get('name', 'Project')}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Client: {project.get('client', '')}    Location: {project.get('location', '')}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(COLUMNS)
    _style_header(ws, header_row)

    subtotal_rows: list[int] = []
    item_no = 0
    for group in boq["groups"]:
        # section header
        ws.append([group["label"]])
        sec_row = ws.max_row
        ws.cell(row=sec_row, column=1).font = BOLD
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=sec_row, column=c).fill = SECTION_FILL

        first = ws.max_row + 1
        for it in group["items"]:
            item_no += 1
            r = ws.max_row + 1
            ws.append([
                item_no, it["description"], it.get("nos"),
                _r(it.get("length_m")), _r(it.get("breadth_m")), _r(it.get("depth_m")),
                it["quantity"], it["unit"], it["rate"], None,
            ])
            # live formula: Amount = Quantity * Rate
            ws.cell(row=r, column=10).value = f"=G{r}*I{r}"
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(row=r, column=c).border = BORDER
        last = ws.max_row

        sub_r = ws.max_row + 1
        ws.append(["", f"Sub-total — {group['label']}", "", "", "", "", "", "", "", None])
        if last >= first:
            ws.cell(row=sub_r, column=10).value = f"=SUM(J{first}:J{last})"
        ws.cell(row=sub_r, column=2).font = BOLD
        ws.cell(row=sub_r, column=10).font = BOLD
        subtotal_rows.append(sub_r)

    gt = ws.max_row + 2
    ws.cell(row=gt, column=2, value="GRAND TOTAL").font = Font(bold=True, size=12)
    if subtotal_rows:
        formula = "=" + "+".join(f"J{r}" for r in subtotal_rows)
        gtcell = ws.cell(row=gt, column=10, value=formula)
        gtcell.font = Font(bold=True, size=12)

    _autosize(ws)
    ws.freeze_panes = f"A{header_row + 1}"


def _bbs_sheet(wb, boq):
    ws = wb.create_sheet("Bar Bending Schedule")
    cols = ["Member", "Bar Mark", "Dia (mm)", "No.", "Cutting Length (m)",
            "Unit Wt (kg/m)", "Total Wt (kg)"]
    ws.append(cols)
    _style_simple_header(ws, cols)
    found = False
    for group in boq["groups"]:
        if group["category"] != "rebar":
            continue
        for it in group["items"]:
            for row in it.get("extra", {}).get("bbs", []):
                found = True
                ws.append([
                    it["description"], row.get("mark"), row.get("dia_mm"),
                    row.get("count"), row.get("cutting_length_m"),
                    row.get("unit_weight_kg_m"), row.get("total_weight_kg"),
                ])
    if not found:
        ws.append(["No reinforcement members yet."])
    _autosize(ws)


def _assumptions_sheet(wb, project, boq):
    ws = wb.create_sheet("Assumptions")
    ws.append(["Assumptions & Basis of Measurement"])
    ws["A1"].font = Font(bold=True, size=13)
    rows = [
        ["Standards", "IS 1200 (measurement), IS 456 (RCC), IS 800/SP6 (steel), SP 34 (detailing)"],
        ["Rebar unit weight", "d^2 / 162 kg/m"],
        ["Lap length (tension)", "50 x dia (configurable)"],
        ["Masonry opening deduction", "openings > 0.1 m2 (IS 1200)"],
        ["Plaster opening deduction", "openings > 0.5 m2 (IS 1200)"],
        ["Reinforcement in concrete", "no deduction"],
        ["Currency", project.get("currency", "INR")],
        ["Note", "AI-assisted draft — quantities must be engineer-verified before use."],
    ]
    for k, v in rows:
        ws.append([k, v])
        ws.cell(row=ws.max_row, column=1).font = BOLD
    if boq.get("errors"):
        ws.append([])
        ws.append(["Unresolved members"])
        ws.cell(row=ws.max_row, column=1).font = BOLD
        for e in boq["errors"]:
            ws.append([e.get("label", ""), e.get("error", "")])
    _autosize(ws)


def _style_simple_header(ws, cols):
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER


def _r(v):
    return round(v, 3) if isinstance(v, (int, float)) else v


def _autosize(ws):
    for col in ws.columns:
        width = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 55))
        ws.column_dimensions[letter].width = width
